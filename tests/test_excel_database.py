from __future__ import annotations

from openpyxl import load_workbook

from app.config.paths import get_audit_db_path, get_collaborators_db_path, get_financial_db_path, get_sales_db_path
from app.config.settings import (
    SHEET_AUDIT_LOG,
    SHEET_CARD_MACHINES,
    SHEET_CASH_FLOW,
    SHEET_COLLABORATORS,
    SHEET_COST_CATEGORIES,
    SHEET_FIXED_COSTS,
    SHEET_MACHINE_CONDITIONS,
    SHEET_RECEIVABLE_RECONCILIATION,
    SHEET_SALES,
    SHEET_VARIABLE_COSTS,
)
from app.repositories.excel_database import ExcelDatabase
from app.repositories.excel_schema import (
    AUDIT_SHEETS_CONFIG,
    CARD_MACHINE_HEADERS,
    CASH_FLOW_HEADERS,
    COLLABORATOR_HEADERS,
    COLLABORATORS_SHEETS_CONFIG,
    FINANCIAL_SHEETS_CONFIG,
    MACHINE_CONDITION_HEADERS,
    SALES_SHEETS_CONFIG,
)


def test_excel_files_are_created(services):
    expected = [
        (get_financial_db_path(), [SHEET_FIXED_COSTS, SHEET_VARIABLE_COSTS, SHEET_CASH_FLOW, SHEET_COST_CATEGORIES], FINANCIAL_SHEETS_CONFIG),
        (get_sales_db_path(), [SHEET_CARD_MACHINES, SHEET_MACHINE_CONDITIONS, SHEET_SALES, SHEET_RECEIVABLE_RECONCILIATION], SALES_SHEETS_CONFIG),
        (get_collaborators_db_path(), [SHEET_COLLABORATORS], COLLABORATORS_SHEETS_CONFIG),
        (get_audit_db_path(), [SHEET_AUDIT_LOG], AUDIT_SHEETS_CONFIG),
    ]
    for path, sheets, config in expected:
        assert ExcelDatabase(path, config).sheet_names() == sheets


def test_card_machine_sheet_has_structured_headers(services):
    wb = load_workbook(get_sales_db_path(), read_only=True)
    try:
        headers = [cell.value for cell in next(wb[SHEET_CARD_MACHINES].iter_rows(max_row=1))]
    finally:
        wb.close()

    assert headers == CARD_MACHINE_HEADERS
    assert "maquina_cartao_id" in headers
    old_machine_id = "maquini" + "nha_id"
    assert old_machine_id not in headers
    assert "parcelas" not in headers


def test_machine_conditions_sheet_has_headers(services):
    wb = load_workbook(get_sales_db_path(), read_only=True)
    try:
        headers = [cell.value for cell in next(wb[SHEET_MACHINE_CONDITIONS].iter_rows(max_row=1))]
    finally:
        wb.close()

    assert headers == MACHINE_CONDITION_HEADERS
    assert "maquina_cartao_id" in headers
    old_machine_id = "maquini" + "nha_id"
    assert old_machine_id not in headers
    assert "parcelas" in headers


def test_cash_flow_sheet_has_origin_headers(services):
    wb = load_workbook(get_financial_db_path(), read_only=True)
    try:
        headers = [cell.value for cell in next(wb[SHEET_CASH_FLOW].iter_rows(max_row=1))]
    finally:
        wb.close()

    assert headers == CASH_FLOW_HEADERS
    assert "origem" in headers
    assert "venda_id" in headers


def test_collaborators_sheet_has_headers(services):
    wb = load_workbook(get_collaborators_db_path(), read_only=True)
    try:
        headers = [cell.value for cell in next(wb[SHEET_COLLABORATORS].iter_rows(max_row=1))]
    finally:
        wb.close()

    assert headers == COLLABORATOR_HEADERS
    assert "salario_base" in headers

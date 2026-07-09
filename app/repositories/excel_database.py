"""Acesso generico a workbooks Excel."""

from __future__ import annotations

import os
import shutil
import uuid
from datetime import datetime, timedelta
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

from app.config.paths import get_backups_dir, get_financial_db_path
from app.repositories.excel_schema import FINANCIAL_SHEETS_CONFIG

BACKUP_TIMESTAMP_FORMAT = "%Y%m%d_%H%M%S"
MAX_BACKUPS_PER_DATABASE = 10
MIN_BACKUP_INTERVAL = timedelta(minutes=10)


class ExcelSaveError(Exception):
    user_message = "Não foi possível salvar. Feche a planilha no Excel e tente novamente."


def _read_sheet_dicts(ws: Worksheet) -> list[dict[str, Any]]:
    if ws.max_row < 1:
        return []
    headers = [str(c.value).strip() for c in ws[1] if c.value is not None]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        if values is None or all(value is None for value in values):
            continue
        rows.append({headers[idx]: value for idx, value in enumerate(values)})
    return rows


HEADER_ALIASES: dict[str, tuple[str, ...]] = {}


def _value_for_header(row: dict[str, Any], header: str) -> Any:
    if header in row and row.get(header) is not None:
        return row.get(header)
    for alias in HEADER_ALIASES.get(header, ()):  # migracao leve de schemas antigos
        if alias in row and row.get(alias) is not None:
            return row.get(alias)
    return ""


def _normalize_row(headers: list[str], row: dict[str, Any]) -> dict[str, Any]:
    return {header: _value_for_header(row, header) for header in headers}


def _write_sheet(ws: Worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for row in rows:
        normalized = _normalize_row(headers, row)
        ws.append([normalized.get(header, "") for header in headers])


def _backup_timestamp(path: str, stem: str) -> datetime | None:
    name = os.path.basename(path)
    prefix = f"{stem}_backup_"
    suffix = ".xlsx"
    if not (name.startswith(prefix) and name.endswith(suffix)):
        return None
    raw = name[len(prefix) : -len(suffix)]
    try:
        return datetime.strptime(raw, BACKUP_TIMESTAMP_FORMAT)
    except ValueError:
        return None


def _list_backups(folder: str, stem: str) -> list[tuple[datetime, str]]:
    if not os.path.isdir(folder):
        return []
    backups: list[tuple[datetime, str]] = []
    for name in os.listdir(folder):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        stamp = _backup_timestamp(path, stem)
        if stamp is not None:
            backups.append((stamp, path))
    backups.sort(key=lambda item: item[0])
    return backups


class ExcelDatabase:
    def __init__(
        self,
        db_path: str | None = None,
        sheets_config: dict[str, list[str]] | None = None,
        backup_dir: str | None = None,
        backup_stem: str | None = None,
    ):
        self.db_path = db_path or get_financial_db_path()
        self.sheets_config = sheets_config or FINANCIAL_SHEETS_CONFIG
        self.backup_dir = backup_dir or get_backups_dir()
        self.backup_stem = backup_stem or os.path.splitext(os.path.basename(self.db_path))[0]

    def create_database(self) -> None:
        os.makedirs(os.path.dirname(self.db_path) or ".", exist_ok=True)
        wb = Workbook()
        first = True
        for sheet_name, headers in self.sheets_config.items():
            ws = wb.active if first else wb.create_sheet(sheet_name)
            ws.title = sheet_name
            ws.append(headers)
            first = False
        self.save_workbook_safe(wb)

    def create_backup(self) -> str | None:
        if not os.path.isfile(self.db_path):
            return None
        os.makedirs(self.backup_dir, exist_ok=True)
        backups = _list_backups(self.backup_dir, self.backup_stem)
        now = datetime.now()
        if backups and now - backups[-1][0] < MIN_BACKUP_INTERVAL:
            self._cleanup_old_backups(backups)
            return None
        stamp = now.strftime(BACKUP_TIMESTAMP_FORMAT)
        dest = os.path.join(self.backup_dir, f"{self.backup_stem}_backup_{stamp}.xlsx")
        shutil.copy2(self.db_path, dest)
        self._cleanup_old_backups(_list_backups(self.backup_dir, self.backup_stem))
        return dest

    def _cleanup_old_backups(self, backups: list[tuple[datetime, str]] | None = None) -> None:
        items = backups if backups is not None else _list_backups(self.backup_dir, self.backup_stem)
        excess = len(items) - MAX_BACKUPS_PER_DATABASE
        if excess <= 0:
            return
        for _stamp, path in items[:excess]:
            try:
                os.remove(path)
            except OSError:
                pass

    def ensure_database(self) -> None:
        if not os.path.isfile(self.db_path):
            self.create_database()
            return
        self.migrate_database()

    def migrate_database(self) -> None:
        wb = self.load_workbook_safe()
        changed = False
        try:
            for sheet_name, headers in self.sheets_config.items():
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                    ws.append(headers)
                    changed = True
                    continue
                ws = wb[sheet_name]
                rows = [_normalize_row(headers, row) for row in _read_sheet_dicts(ws)]
                current_headers = [str(c.value).strip() for c in ws[1] if c.value is not None]
                if current_headers != headers:
                    _write_sheet(ws, headers, rows)
                    changed = True
            if changed:
                self.create_backup()
                self.save_workbook_safe(wb)
            else:
                wb.close()
        except Exception:
            wb.close()
            raise

    def load_workbook_safe(self) -> WorkbookType:
        return load_workbook(self.db_path, read_only=False, data_only=False)

    def save_workbook_safe(self, wb: WorkbookType) -> None:
        folder = os.path.dirname(self.db_path) or "."
        os.makedirs(folder, exist_ok=True)
        tmp = os.path.join(folder, f".{os.path.basename(self.db_path)}.{uuid.uuid4().hex}.tmp.xlsx")
        try:
            wb.save(tmp)
            wb.close()
            os.replace(tmp, self.db_path)
        except (PermissionError, OSError) as exc:
            if os.path.isfile(tmp):
                os.remove(tmp)
            raise ExcelSaveError() from exc

    def read_sheet(self, sheet_name: str) -> list[dict[str, Any]]:
        self.ensure_database()
        wb = self.load_workbook_safe()
        try:
            if sheet_name not in wb.sheetnames:
                return []
            return _read_sheet_dicts(wb[sheet_name])
        finally:
            wb.close()

    def read_sheet_dataframe(self, sheet_name: str) -> pd.DataFrame:
        return pd.DataFrame(self.read_sheet(sheet_name))

    def write_sheet(self, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
        self.ensure_database()
        wb = self.load_workbook_safe()
        try:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            _write_sheet(wb[sheet_name], headers, rows)
            self.save_workbook_safe(wb)
            self.create_backup()
        except Exception:
            wb.close()
            raise

    def append_row(self, sheet_name: str, headers: list[str], row: dict[str, Any]) -> dict[str, Any]:
        rows = self.read_sheet(sheet_name)
        normalized = _normalize_row(headers, row)
        rows.append(normalized)
        self.write_sheet(sheet_name, headers, rows)
        return normalized

    def update_row(self, sheet_name: str, headers: list[str], id_field: str, item_id: str, changes: dict) -> dict[str, Any]:
        rows = self.read_sheet(sheet_name)
        updated: dict[str, Any] | None = None
        for row in rows:
            if str(row.get(id_field, "")) == str(item_id):
                row.update(changes)
                updated = _normalize_row(headers, row)
                break
        if updated is None:
            raise KeyError("Registro não encontrado.")
        self.write_sheet(sheet_name, headers, rows)
        return updated

    def sheet_names(self) -> list[str]:
        self.ensure_database()
        wb = self.load_workbook_safe()
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()



